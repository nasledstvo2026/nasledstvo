#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Счёт №74"

thin = Side(style='thin')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
bold = Font(bold=True)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

# Шапка
ws.merge_cells('A1:F1')
ws['A1'] = 'Счет на оплату № 74 от 19 мая 2026 г.'
ws['A1'].font = Font(bold=True, size=14)
ws['A1'].alignment = Alignment(horizontal='center')

ws.merge_cells('A3:B3')
ws['A3'] = 'Поставщик:'
ws['A3'].font = bold
ws.merge_cells('C3:F3')
ws['C3'] = 'ООО "РАДИОДОМ", ИНН 7802630056, КПП 780201001, 195277, г. Санкт-Петербург, ул. Гельсингфорсская, д. 3, лит. З, оф. 407, тел.: 3330365'

ws.merge_cells('A4:B4')
ws['A4'] = 'Покупатель:'
ws['A4'].font = bold
ws.merge_cells('C4:F4')
ws['C4'] = 'ООО "СТК", ИНН 7720427871, КПП 773601001, 117312, г. Москва, ул. Вавилова, д. 23, стр. 10'

# Заголовки таблицы
row = 6
headers = ['№', 'Товары (работы, услуги)', 'Кол-во', 'Ед.', 'Цена', 'Сумма']
col_widths = [5, 55, 10, 8, 15, 15]
for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    c = ws.cell(row=row, column=i, value=h)
    c.font = bold
    c.border = border
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center')
    ws.column_dimensions[get_column_letter(i)].width = w

# Данные
items = [
    [1, 'Стационарная радиостанция LIRA DR-1000V (new case)', 1, 'шт', 82000.00, 82000.00],
    [2, 'Радиостанция автомобильная Lira DM-1000V DMR', 3, 'шт', 32000.00, 96000.00],
    [3, 'Цифровая радиостанция Lira DP-2000V DMR', 3, 'шт', 18000.00, 54000.00],
    [4, 'Цифровая радиостанция Lira DP-2600V DMR', 1, 'шт', 18500.00, 18500.00],
    [5, 'Аккумуляторная батарея B-2000USB', 4, 'шт', 3400.00, 13600.00],
    [6, 'Кабель программирования для DR-1000 (new case)', 1, 'шт', 2500.00, 2500.00],
    [7, 'Кабель программирования для DM-1000, разъем подключения к компьютеру: USB', 1, 'шт', 2500.00, 2500.00],
    [8, 'Кабель программирования для радиостанций Lira DP-2000', 1, 'шт', 2500.00, 2500.00],
    [9, 'Антенна автомобильная на магнитном основании AW-6V', 3, 'компл', 5200.00, 15600.00],
    [10, 'Антенна базовая Project X-30V', 1, 'шт', 7200.00, 7200.00],
    [11, 'Разъем U-113/5D', 2, 'шт', 500.00, 1000.00],
    [12, 'Кабель 5D-FB peeg 15 м', 1, 'м', 140.00, 2100.00],
]

for i, item in enumerate(items):
    r = row + 1 + i
    for j, val in enumerate(item, 1):
        c = ws.cell(row=r, column=j, value=val)
        c.border = border
        if j in (5, 6):
            c.number_format = '# ##0.00'
            c.alignment = Alignment(horizontal='right')
        elif j == 1:
            c.alignment = Alignment(horizontal='center')
        elif j == 3:
            c.alignment = Alignment(horizontal='center')
        elif j == 4:
            c.alignment = Alignment(horizontal='center')

# Итого
r = row + 1 + len(items)
ws.merge_cells(f'A{r}:E{r}')
ws.cell(row=r, column=1, value='Итого:').font = bold
ws.cell(row=r, column=1).border = border
ws.cell(row=r, column=1).alignment = Alignment(horizontal='right')
for col in range(2, 6):
    ws.cell(row=r, column=col).border = border
c = ws.cell(row=r, column=6, value=297500.00)
c.font = bold
c.border = border
c.number_format = '# ##0.00'
c.alignment = Alignment(horizontal='right')

r += 1
ws.merge_cells(f'A{r}:E{r}')
ws.cell(row=r, column=1, value='В том числе НДС 22%:').border = border
ws.cell(row=r, column=1).alignment = Alignment(horizontal='right')
for col in range(2, 6):
    ws.cell(row=r, column=col).border = border
c = ws.cell(row=r, column=6, value=53647.55)
c.border = border
c.number_format = '# ##0.00'
c.alignment = Alignment(horizontal='right')

r += 1
ws.merge_cells(f'A{r}:E{r}')
ws.cell(row=r, column=1, value='Всего к оплате:').font = bold
ws.cell(row=r, column=1).border = border
ws.cell(row=r, column=1).alignment = Alignment(horizontal='right')
for col in range(2, 6):
    ws.cell(row=r, column=col).border = border
c = ws.cell(row=r, column=6, value=297500.00)
c.font = bold
c.border = border
c.number_format = '# ##0.00'
c.alignment = Alignment(horizontal='right')

r += 2
ws.merge_cells(f'A{r}:F{r}')
ws.cell(row=r, column=1, value='Всего наименований 12, на сумму 297 500,00 руб.')
r += 1
ws.merge_cells(f'A{r}:F{r}')
ws.cell(row=r, column=1, value='Двести девяносто семь тысяч пятьсот рублей 00 копеек')
r += 1
ws.merge_cells(f'A{r}:F{r}')
ws.cell(row=r, column=1, value='Оплатить не позднее 22.05.2026')
ws.cell(row=r, column=1).font = bold

r += 2
ws.merge_cells(f'A{r}:C{r}')
ws.cell(row=r, column=1, value='Руководитель Антошкин О. П.')
ws.merge_cells(f'D{r}:F{r}')
ws.cell(row=r, column=4, value='Бухгалтер Сорокина О. Н.')

out = '/home/user1/.openclaw/workspace/Счёт_74_от_19.05.2026.xlsx'
wb.save(out)
print(f'Saved: {out}')
